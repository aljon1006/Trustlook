import requests
from requests import Session
import json
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook, load_workbook

address = "rB67WCo91hbzz8hkc3oDgnjDckSvekeZq8"
# url =f"https://api.xrpscan.com/api/v1/account/{address}"
url = "https://api.xrpscan.com/api/v1/names/well-known"
# url = "https://api.xrpscan.com/api/v1/account/rB67WCo91hbzz8hkc3oDgnjDckSvekeZq8"


headers = {
  'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36',
}
session = requests.Session()
session.trust_env = False
response = session.get(url, headers=headers)
print(response)
data = json.loads(response.text)

filename = "xrpscan.xlsx"
column_names = ['name', 'address', 'desc', 'twitter', 'domain', 'verified']
wb = load_workbook(filename=filename, read_only=False)
ws = wb.sheetnames
for sheet in ws:
    std = wb[sheet]
    if sheet == 'data':
        wb.remove(std)
new_ws = wb.create_sheet('data')
new_ws.append(column_names)

for item in data:
    name = "N/A"
    if "name" in item:
        name = item["name"]
    address = "N/A"
    if "account" in item:
        address = item["account"]
    desc = "N/A"
    if "desc" in item:
        desc = item["desc"]
    twitter = "N/A"
    if "twitter" in item:
        twitter = item[ "twitter" ]
    domain = "N/A"
    if "domain" in item:
        domain = item[ "domain" ]
    verified = "N/A"
    if "verified" in item:
        verified = item[ "verified" ]

    new_ws.append([name, address, desc, twitter, domain, verified])
    print(name, address, domain)

wb.save(filename=filename)

root = Tk()
root.withdraw()
messagebox.showinfo('Python Alert', message=f'Completed processing: xrpscan.py. See file for result: {filename}')
root.mainloop()

