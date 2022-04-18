import json
import pandas as pd
from openpyxl import Workbook, load_workbook
from tkinter import *
from tkinter import messagebox
import requests
from playsound import playsound
import time


def check_new_TL():
    issuer = ''

    print('Searching for new TLs...')
    count = 0
    filename = 'issuedcurrencies.xlsx'
    column_names = ['Currency', 'Address','Limit','Trustlines']
    currency_list = []
    df_currency_list = pd.read_excel(io = filename,sheet_name='Existing_All')

    # get current list from excel file
    for index, row in df_currency_list.iterrows():
        currency_list.append(row['Address'])

    # prepare excel file
    wb = load_workbook(filename=filename, read_only=False)
    ws = wb.sheetnames
    for sheet in ws:
        std = wb[sheet]
        if sheet == 'NEW':
            wb.remove(std)
    new_ws = wb.create_sheet('NEW')
    new_ws.append(column_names)

    # get data from API
    headers = {
      'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36',
    }
    url = 'https://xrpldata.com/api/v1/tokens'
    session = requests.Session()
    session.trust_env = False
    response = session.get(url, headers=headers)
    data = json.loads(response.text)
    issuers = data['issuers']

    # get all currencies from API and check against existing list; Add to excel file if new
    for issuer in issuers:
        currency = data['issuers'][issuer]
        if currency.get('tokens'):
            tokens = currency[ 'tokens' ]
            currency = tokens[0]['currency']
            if not (issuer in currency_list):
                limit = tokens[0]['amount']
                trustlines = tokens[0]['trustlines']
                new_ws.append([currency, issuer, limit, trustlines])
                print(currency, issuer, limit, trustlines)
                print(f'https://xumm.community/?issuer={issuer}&currency={currency}&limit={limit}')
                print()
                count += 1


    wb.save(filename=filename)
    # df_new = pd.read_excel(io = filename,sheet_name='NEW')
    # print(df_new)

    if count:
        playsound(r'D:\Programming\Alarm\alarm.mp3')
        root = Tk()
        root.withdraw()
        messagebox.showinfo('Python Alert', message='NEW TRUSTLINE ALERT')
        root.mainloop()
    else:
        print('No new TLs found. Next run after 3 minutes')
        print()

while True:
    check_new_TL()
    time.sleep(180)

