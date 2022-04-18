import requests
import json
import pandas as pd
from openpyxl import Workbook, load_workbook
import xlwings as xw
import time

headers = {
  'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36',
}

filename = 'XUMM Balances.xlsx'
wallets = []
address_list = pd.read_excel(filename)
column_names = ['Currency','Balance','XRPBalance']

row_number = 2
total_balance = 0
total_USD = 0
total_PHP = 0

wb = load_workbook(filename=filename, read_only=False)

for index, row in address_list.iterrows():
    totalXRP = float(0)
    address = row['Address']
    account_name = row['Account']
    print(account_name)

    # get balances
    url = f'http://data.ripple.com/v2/accounts/{address}/balances'
    print(url)
    session = requests.Session()
    session.trust_env = False
    response = session.get(url, headers=headers)
    data = json.loads(response.text)
    balances = data['balances']

    xrpValue = 0
    values = []

    ws = wb.sheetnames
    for sheet in ws:
        std = wb[ sheet ]
        if sheet == account_name:
            wb.remove(std)
    new_ws = wb.create_sheet(account_name)
    new_ws.append(column_names)

    print(balances)

    for balance in balances:
        currency = balance['currency']
        isXRP = False
        try:
            issuer = balance['counterparty']
        except:
            isXRP = True
            xrpValue = float(balance[ 'value' ])
        value = float(balance['value'])
        print(currency)

        if not isXRP:
            convert = f'http://data.ripple.com/v2/exchange_rates/XRP/{currency}+{issuer}'
            print(convert)
            session = requests.Session()
            session.trust_env = False
            response = session.get(convert, headers=headers)
            data = json.loads(response.text)
            retry_count = 0

            rate = float(data['rate'])

            try:
                xrpValue = float(value) / rate
            except:
                xrpValue = 0
        values.append(float(xrpValue))
        new_ws.append([ currency, value, xrpValue ])
        totalXRP += xrpValue
        time.sleep(0.5)

    main_ws = wb[ 'Sheet1' ]

    usd_convert = 'http://data.ripple.com/v2/exchange_rates/XRP/USD+rvYAfWj5gh67oV6fW32ZzP3Aw4Eubs59B'
    session = requests.Session()
    session.trust_env = False
    response = session.get(usd_convert, headers=headers)
    data = json.loads(response.text)
    usd_rate = float(data['rate'])

    php_convert = 'http://api.coingecko.com/api/v3/simple/price?ids=ripple&vs_currencies=php'
    session = requests.Session()
    session.trust_env = False
    response = session.get(php_convert, headers=headers)
    data = json.loads(response.text)
    # print(data['ripple']['php'])
    php_rate = float(data['ripple']['php'])

    account_usd = totalXRP * usd_rate
    account_php = totalXRP * php_rate

    main_ws[f'C{row_number}'] = totalXRP
    main_ws[f'D{row_number}'] = account_usd
    main_ws[f'E{row_number}'] = account_php

    total_balance += totalXRP
    total_USD += account_usd
    total_PHP += account_php

    row_number += 1

    print()

    time.sleep(3)
# wb = load_workbook(filename=filename, read_only=False)
main_ws = wb[ 'Sheet1' ]
main_ws[f'C{row_number}'] = total_balance
main_ws[f'D{row_number}'] = total_USD
main_ws[f'E{row_number}'] = total_PHP

wb.save(filename=filename)

from tkinter import *
from tkinter import messagebox

root = Tk()
root.withdraw()
messagebox.showinfo('Python Alert', message='Completed')
# root.mainloop()